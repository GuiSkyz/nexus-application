import io

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from app.main import app


@pytest.mark.asyncio
async def test_operational_report_summary_and_exports():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        await client.post(
            "/api/v1/auth/login",
            json={"email": "coordenador@nexusops.com", "password": "senha123"},
        )
        summary_response = await client.get("/api/v1/reports/operational")
        xlsx_response = await client.get("/api/v1/reports/operational.xlsx")
        pdf_response = await client.get("/api/v1/reports/operational.pdf")

    assert summary_response.status_code == 200
    summary = summary_response.json()
    assert summary["overall"]["inspections"] > 0
    assert len(summary["teams"]) >= 3
    assert len(summary["fleet"]) >= 3

    assert xlsx_response.status_code == 200
    assert xlsx_response.content.startswith(b"PK")
    assert "spreadsheetml" in xlsx_response.headers["content-type"]
    workbook = load_workbook(io.BytesIO(xlsx_response.content))
    assert workbook.sheetnames == ["Resumo Operacional", "Dados"]
    assert workbook["Resumo Operacional"]["A1"].value == "NexusOps · Relatório Operacional"
    assert len(workbook["Resumo Operacional"]._charts) == 1

    assert pdf_response.status_code == 200
    assert pdf_response.content.startswith(b"%PDF")
    assert pdf_response.headers["content-type"] == "application/pdf"
