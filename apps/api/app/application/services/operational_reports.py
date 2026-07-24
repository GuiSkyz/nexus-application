import io
from dataclasses import dataclass
from datetime import date
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

NAVY = "061F4A"
BLUE = "0757C8"
CYAN = "00B8E6"
SURFACE = "F5F7FA"
TEXT = "17202E"


@dataclass(frozen=True)
class OperationalRecord:
    occurred_on: date
    team: str
    vehicle: str
    vehicle_plate: str
    inspections: int
    checklist_items: int
    conforming_items: int
    nonconformities: int
    critical_nonconformities: int
    resolved_nonconformities: int

    @property
    def compliance_rate(self) -> float:
        if self.checklist_items == 0:
            return 0
        return self.conforming_items / self.checklist_items


SAMPLE_OPERATIONAL_RECORDS = [
    OperationalRecord(date(2026, 7, 20), "Equipe Alfa", "Caminhonete 12", "ABC1D23", 18, 216, 204, 12, 1, 9),
    OperationalRecord(date(2026, 7, 20), "Equipe Beta", "Caminhonete 08", "DEF4G56", 15, 180, 166, 14, 2, 10),
    OperationalRecord(date(2026, 7, 21), "Equipe Alfa", "Caminhonete 12", "ABC1D23", 20, 240, 231, 9, 0, 8),
    OperationalRecord(date(2026, 7, 21), "Equipe Gama", "Van 04", "GHI7J89", 13, 156, 142, 14, 1, 11),
    OperationalRecord(date(2026, 7, 22), "Equipe Beta", "Caminhonete 08", "DEF4G56", 17, 204, 194, 10, 1, 9),
    OperationalRecord(date(2026, 7, 22), "Equipe Gama", "Van 04", "GHI7J89", 16, 192, 181, 11, 0, 9),
    OperationalRecord(date(2026, 7, 23), "Equipe Alfa", "Caminhonete 12", "ABC1D23", 19, 228, 220, 8, 1, 6),
    OperationalRecord(date(2026, 7, 23), "Equipe Beta", "Caminhonete 08", "DEF4G56", 18, 216, 207, 9, 0, 7),
]


def filter_records(
    records: Iterable[OperationalRecord],
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[OperationalRecord]:
    return [
        record
        for record in records
        if (start_date is None or record.occurred_on >= start_date)
        and (end_date is None or record.occurred_on <= end_date)
    ]


def _aggregate(records: list[OperationalRecord], key: str) -> list[dict]:
    groups: dict[str, dict] = {}
    for record in records:
        group_key = getattr(record, key)
        group = groups.setdefault(
            group_key,
            {
                "name": group_key,
                "inspections": 0,
                "checklistItems": 0,
                "conformingItems": 0,
                "nonconformities": 0,
                "criticalNonconformities": 0,
                "resolvedNonconformities": 0,
            },
        )
        group["inspections"] += record.inspections
        group["checklistItems"] += record.checklist_items
        group["conformingItems"] += record.conforming_items
        group["nonconformities"] += record.nonconformities
        group["criticalNonconformities"] += record.critical_nonconformities
        group["resolvedNonconformities"] += record.resolved_nonconformities

    for group in groups.values():
        items = group["checklistItems"]
        group["complianceRate"] = (
            round(group["conformingItems"] / items * 100, 1) if items else 0
        )
    return sorted(groups.values(), key=lambda item: item["complianceRate"], reverse=True)


def build_operational_summary(records: list[OperationalRecord]) -> dict:
    checklist_items = sum(record.checklist_items for record in records)
    conforming_items = sum(record.conforming_items for record in records)
    nonconformities = sum(record.nonconformities for record in records)
    resolved = sum(record.resolved_nonconformities for record in records)

    return {
        "period": {
            "start": min((record.occurred_on for record in records), default=None),
            "end": max((record.occurred_on for record in records), default=None),
        },
        "overall": {
            "inspections": sum(record.inspections for record in records),
            "complianceRate": round(conforming_items / checklist_items * 100, 1)
            if checklist_items
            else 0,
            "nonconformities": nonconformities,
            "criticalNonconformities": sum(
                record.critical_nonconformities for record in records
            ),
            "resolutionRate": round(resolved / nonconformities * 100, 1)
            if nonconformities
            else 0,
        },
        "teams": _aggregate(records, "team"),
        "fleet": _aggregate(records, "vehicle"),
    }


def generate_operational_xlsx(records: list[OperationalRecord]) -> bytes:
    summary = build_operational_summary(records)
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Resumo Operacional"
    details = workbook.create_sheet("Dados")

    dashboard.sheet_view.showGridLines = False
    dashboard.merge_cells("A1:F2")
    dashboard["A1"] = "NexusOps · Relatório Operacional"
    dashboard["A1"].fill = PatternFill("solid", fgColor=NAVY)
    dashboard["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    dashboard["A1"].alignment = Alignment(vertical="center")

    dashboard["A4"] = "Indicador"
    dashboard["B4"] = "Resultado"
    metrics = [
        ("Inspeções", summary["overall"]["inspections"], "#,##0"),
        ("Conformidade", summary["overall"]["complianceRate"] / 100, "0.0%"),
        ("Não conformidades", summary["overall"]["nonconformities"], "#,##0"),
        ("NCs críticas", summary["overall"]["criticalNonconformities"], "#,##0"),
        ("Taxa de resolução", summary["overall"]["resolutionRate"] / 100, "0.0%"),
    ]
    for row_index, (label, value, number_format) in enumerate(metrics, start=5):
        dashboard.cell(row=row_index, column=1, value=label)
        dashboard.cell(row=row_index, column=2, value=value)
        dashboard.cell(row=row_index, column=2).number_format = number_format

    team_start = 4
    dashboard.cell(row=team_start, column=4, value="Equipe")
    dashboard.cell(row=team_start, column=5, value="Inspeções")
    dashboard.cell(row=team_start, column=6, value="Conformidade")
    for row_index, team in enumerate(summary["teams"], start=team_start + 1):
        dashboard.cell(row=row_index, column=4, value=team["name"])
        dashboard.cell(row=row_index, column=5, value=team["inspections"])
        dashboard.cell(row=row_index, column=6, value=team["complianceRate"] / 100)
        dashboard.cell(row=row_index, column=6).number_format = "0.0%"

    fleet_start = 12
    dashboard.cell(row=fleet_start, column=1, value="Frota")
    dashboard.cell(row=fleet_start, column=2, value="Inspeções")
    dashboard.cell(row=fleet_start, column=3, value="Conformidade")
    dashboard.cell(row=fleet_start, column=4, value="NCs")
    for row_index, vehicle in enumerate(summary["fleet"], start=fleet_start + 1):
        dashboard.cell(row=row_index, column=1, value=vehicle["name"])
        dashboard.cell(row=row_index, column=2, value=vehicle["inspections"])
        dashboard.cell(row=row_index, column=3, value=vehicle["complianceRate"] / 100)
        dashboard.cell(row=row_index, column=3).number_format = "0.0%"
        dashboard.cell(row=row_index, column=4, value=vehicle["nonconformities"])

    header_fill = PatternFill("solid", fgColor=BLUE)
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        bottom=Side(style="thin", color="D8E0E8")
    )
    for row in (4, fleet_start):
        for cell in dashboard[row]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left")
    for row in dashboard.iter_rows(min_row=5, max_row=dashboard.max_row):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Conformidade por equipe"
    chart.y_axis.title = "Equipe"
    chart.x_axis.title = "Conformidade"
    chart.height = 6
    chart.width = 10
    data = Reference(
        dashboard,
        min_col=6,
        min_row=team_start,
        max_row=team_start + len(summary["teams"]),
    )
    categories = Reference(
        dashboard,
        min_col=4,
        min_row=team_start + 1,
        max_row=team_start + len(summary["teams"]),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    dashboard.add_chart(chart, "H4")

    detail_headers = [
        "Data",
        "Equipe",
        "Veículo",
        "Placa",
        "Inspeções",
        "Itens",
        "Conformes",
        "Conformidade",
        "Não conformidades",
        "Críticas",
        "Resolvidas",
    ]
    details.append(detail_headers)
    for record in records:
        details.append(
            [
                record.occurred_on,
                record.team,
                record.vehicle,
                record.vehicle_plate,
                record.inspections,
                record.checklist_items,
                record.conforming_items,
                record.compliance_rate,
                record.nonconformities,
                record.critical_nonconformities,
                record.resolved_nonconformities,
            ]
        )
    for cell in details[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in details["H"][1:]:
        cell.number_format = "0.0%"
    for cell in details["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions

    widths = {
        "A": 22,
        "B": 16,
        "C": 18,
        "D": 18,
        "E": 14,
        "F": 16,
    }
    for column, width in widths.items():
        dashboard.column_dimensions[column].width = width
    for column_index, header in enumerate(detail_headers, start=1):
        details.column_dimensions[get_column_letter(column_index)].width = max(
            12, min(24, len(header) + 3)
        )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_operational_pdf(records: list[OperationalRecord]) -> bytes:
    summary = build_operational_summary(records)
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Relatório operacional NexusOps",
        author="NexusOps",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "TitleNexus",
        parent=styles["Title"],
        textColor=colors.HexColor(f"#{NAVY}"),
        fontName="Helvetica-Bold",
        fontSize=18,
    )
    heading = ParagraphStyle(
        "HeadingNexus",
        parent=styles["Heading2"],
        textColor=colors.HexColor(f"#{NAVY}"),
        fontName="Helvetica-Bold",
        fontSize=11,
    )
    story = [
        Paragraph("NexusOps · Relatório Operacional", title),
        Paragraph(
            "Conformidade de equipes e frota no período selecionado.",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]

    metric_table = Table(
        [
            ["Inspeções", "Conformidade", "Não conformidades", "Críticas", "Resolução"],
            [
                summary["overall"]["inspections"],
                f'{summary["overall"]["complianceRate"]:.1f}%',
                summary["overall"]["nonconformities"],
                summary["overall"]["criticalNonconformities"],
                f'{summary["overall"]["resolutionRate"]:.1f}%',
            ],
        ],
        colWidths=[48 * mm] * 5,
    )
    metric_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor(f"#{SURFACE}")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8E0E8")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.extend([metric_table, Spacer(1, 6 * mm), Paragraph("Por equipe", heading)])

    team_rows = [["Equipe", "Inspeções", "Conformidade", "NCs", "Críticas", "Resolvidas"]]
    for team in summary["teams"]:
        team_rows.append(
            [
                team["name"],
                team["inspections"],
                f'{team["complianceRate"]:.1f}%',
                team["nonconformities"],
                team["criticalNonconformities"],
                team["resolvedNonconformities"],
            ]
        )
    team_table = Table(team_rows, repeatRows=1, colWidths=[58 * mm, 32 * mm, 36 * mm, 28 * mm, 28 * mm, 32 * mm])
    team_table.setStyle(_report_table_style())
    story.extend([team_table, Spacer(1, 6 * mm), Paragraph("Por veículo", heading)])

    fleet_rows = [["Veículo", "Inspeções", "Conformidade", "NCs", "Críticas", "Resolvidas"]]
    for vehicle in summary["fleet"]:
        fleet_rows.append(
            [
                vehicle["name"],
                vehicle["inspections"],
                f'{vehicle["complianceRate"]:.1f}%',
                vehicle["nonconformities"],
                vehicle["criticalNonconformities"],
                vehicle["resolvedNonconformities"],
            ]
        )
    fleet_table = Table(fleet_rows, repeatRows=1, colWidths=[58 * mm, 32 * mm, 36 * mm, 28 * mm, 28 * mm, 32 * mm])
    fleet_table.setStyle(_report_table_style())
    story.append(fleet_table)

    document.build(story)
    return output.getvalue()


def _report_table_style() -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BLUE}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{SURFACE}")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8E0E8")),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]
    )
